import os
import sys
import json
import base64
import re
import uuid
from datetime import datetime

from flask import Flask, render_template, request, jsonify, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv
from werkzeug.utils import secure_filename

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSHEETS_AVAILABLE = True
except ImportError:
    GSHEETS_AVAILABLE = False

load_dotenv()

# ── Path setup (supports PyInstaller / pywebview) ──
if getattr(sys, "frozen", False):
    _base_dir = os.path.dirname(sys.executable)
    _bundle_dir = sys._MEIPASS
else:
    _base_dir = os.path.dirname(os.path.abspath(__file__))
    _bundle_dir = _base_dir

app = Flask(
    __name__,
    template_folder=os.path.join(_bundle_dir, "templates"),
    static_folder=os.path.join(_bundle_dir, "static"),
)
app.config["UPLOAD_FOLDER"] = os.path.join(_base_dir, "uploads")
app.config["EXPORT_FOLDER"] = os.path.join(_base_dir, "exports")
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024

ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "webp", "gif", "pdf"}
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
os.makedirs(app.config["EXPORT_FOLDER"], exist_ok=True)

# client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Google Sheets connection (persists while app is running)
sheets_state: dict = {"connected": False, "spreadsheet_id": None, "sheet_url": None}

SHEET_HEADERS = ["Timestamp", "Expense/Income", "Date", "Month", "Amount",
                 "Type of Expense", "Description"]


# Helpers

def allowed_file(filename: str) -> bool:
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def encode_image(image_path: str) -> str:
    with open(image_path, "rb") as f:
        return base64.b64encode(f.read()).decode("utf-8")


def extract_text_from_pdf(pdf_path: str) -> str:
    """Extract text from PDF file using pypdf."""
    try:
        from pypdf import PdfReader
        reader = PdfReader(pdf_path)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"Error extracting PDF text: {e}")
        return ""


# receipt parsing
# (AI parsing temporarily disabled)
# TEXT_EXTRACTION_PROMPT = """Extract all visible text from this receipt image. Return plain text only, preserving line breaks and structure."""
#
# PARSE_PROMPT = """Analyze this receipt image and extract the following information as JSON.
# Be as accurate as possible. If a field is not visible, use null.
#
# Return ONLY valid JSON with this exact structure:
# {
#   "store_name": "Name of the store/business",
#   "date": "YYYY-MM-DD format or null",
#   "items": [
#     {
#       "name": "Item description",
#       "quantity": 1,
#       "price": 0.00,
#       "category": "one of: Food & Groceries, Dining & Restaurants, Transportation, Healthcare, Entertainment, Shopping, Utilities, Office Supplies, Personal Care, Other"
#     }
#   ],
#   "subtotal": 0.00,
#   "tax": 0.00,
#   "total": 0.00,
#   "payment_method": "Cash/Credit/Debit/Other or null"
# }
# """
#
# def extract_text_from_image(image_path: str) -> str:
#     """Extract raw text from receipt image using Vision API."""
#     b64 = encode_image(image_path)
#     ext = image_path.rsplit(".", 1)[1].lower()
#     mime = f"image/{'jpeg' if ext in ('jpg', 'jpeg') else ext}"
#
#     try:
#         response = client.chat.completions.create(
#             model="gpt-4o",
#             messages=[{
#                 "role": "user",
#                 "content": [
#                     {"type": "text", "text": TEXT_EXTRACTION_PROMPT},
#                     {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
#                 ],
#             }],
#             max_tokens=1000,
#         )
#         return response.choices[0].message.content.strip()
#     except Exception:
#         return ""


def try_pattern_match(text: str) -> dict | None:
    """Attempt to extract receipt data using pattern matching. Returns None if insufficient data found."""
    data = {
        "store_name": None,
        "date": None,
        "items": [],
        "subtotal": None,
        "tax": None,
        "total": None,
        "payment_method": None,
    }
    
    lines = text.split('\n')
    
    # Try to find date (YYYY-MM-DD, MM/DD/YY, MM/DD/YYYY, etc.)
    date_patterns = [
        r'(\d{4}-\d{2}-\d{2})',
        r'(\d{1,2}/\d{1,2}/\d{2,4})',
    ]
    for pattern in date_patterns:
        match = re.search(pattern, text)
        if match:
            date_str = match.group(1)
            # Try to normalize to YYYY-MM-DD
            try:
                if '-' in date_str:
                    data["date"] = date_str
                elif '/' in date_str:
                    parts = date_str.split('/')
                    if len(parts[2]) == 2:
                        year = int(parts[2]) + (2000 if int(parts[2]) < 50 else 1900)
                    else:
                        year = int(parts[2])
                    data["date"] = f"{year:04d}-{int(parts[0]):02d}-{int(parts[1]):02d}"
            except:
                pass
            if data["date"]:
                break
    
    # Try to find total (look for "Total" or "TOTAL" followed by $ or number)
    total_patterns = [
        r'(?:Total|TOTAL|Total Due|TOTAL DUE)[:\s]*\$?\s*(\d+[.,]\d{2})',
        r'(?:Grand Total|GRAND TOTAL)[:\s]*\$?\s*(\d+[.,]\d{2})',
    ]
    for pattern in total_patterns:
        match = re.search(pattern, text, re.IGNORECASE)
        if match:
            try:
                amount_str = match.group(1).replace(',', '.')
                data["total"] = float(amount_str)
                break
            except:
                pass
    
    # Try to find subtotal and tax
    subtotal_match = re.search(r'(?:Subtotal|SUBTOTAL)[:\s]*\$?\s*(\d+[.,]\d{2})', text, re.IGNORECASE)
    if subtotal_match:
        try:
            data["subtotal"] = float(subtotal_match.group(1).replace(',', '.'))
        except:
            pass
    
    tax_match = re.search(r'(?:Tax|TAX|Sales Tax)[:\s]*\$?\s*(\d+[.,]\d{2})', text, re.IGNORECASE)
    if tax_match:
        try:
            data["tax"] = float(tax_match.group(1).replace(',', '.'))
        except:
            pass
    
    # Try to find store name (usually in first few lines and capitalized)
    for line in lines[:5]:
        line_clean = line.strip()
        if line_clean and len(line_clean) > 3 and not any(c.isdigit() for c in line_clean[:5]):
            if data["store_name"] is None or len(line_clean) < len(data["store_name"]):
                data["store_name"] = line_clean
    
    # Simple item extraction - look for price patterns
    price_pattern = r'\$?\s*(\d+[.,]\d{2})\s*$'
    for line in lines:
        if re.search(price_pattern, line.strip()):
            item_match = re.match(r'(.+?)\s+\$?\s*(\d+[.,]\d{2})\s*$', line.strip())
            if item_match:
                try:
                    data["items"].append({
                        "name": item_match.group(1).strip(),
                        "quantity": 1,
                        "price": float(item_match.group(2).replace(',', '.')),
                        "category": "Other",
                    })
                except:
                    pass
    
    # Check if we have meaningful data
    has_enough_data = data["total"] is not None and (data["store_name"] or data["date"])
    
    return data if has_enough_data else None


def parse_receipt(file_path: str) -> dict:
    """Parse receipt using pattern matching. Supports images and PDFs."""
    # Extract text based on file type
    ext = file_path.rsplit(".", 1)[1].lower()
    text = ""
    
    if ext == "pdf":
        text = extract_text_from_pdf(file_path)
    else:
        # For images, return template (OCR would be needed here)
        text = ""
    
    # Try pattern matching if we have text
    if text:
        matched_data = try_pattern_match(text)
        if matched_data:
            return matched_data
    
    # Return empty template if no data found
    return {
        "store_name": None,
        "date": None,
        "items": [],
        "subtotal": None,
        "tax": None,
        "total": None,
        "payment_method": None,
    }

def receipt_to_record(parsed: dict) -> dict:
    """Convert parsed receipt JSON into a flat ledger record."""
    receipt_date = parsed.get("date") or ""
    month = ""
    if receipt_date:
        try:
            month = datetime.strptime(receipt_date, "%Y-%m-%d").strftime("%B %Y")
        except ValueError:
            pass

    # Dominant category from items
    cat_totals: dict[str, float] = {}
    for item in parsed.get("items", []):
        cat = item.get("category", "Other")
        cat_totals[cat] = cat_totals.get(cat, 0) + (
            item.get("price", 0) * item.get("quantity", 1)
        )
    expense_type = max(cat_totals, key=cat_totals.get) if cat_totals else "Other"

    # Description = store + first few items
    store = parsed.get("store_name") or "Unknown Store"
    names = [i.get("name", "") for i in parsed.get("items", []) if i.get("name")]
    desc = store
    if names:
        desc += " — " + ", ".join(names[:5])
        if len(names) > 5:
            desc += f", +{len(names) - 5} more"

    return {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "expense_income": "Expense",
        "date": receipt_date,
        "month": month,
        "amount": parsed.get("total") or 0,
        "type_of_expense": expense_type,
        "description": desc,
    }


# Google Sheets helpers
def _extract_spreadsheet_id(url: str) -> str | None:
    m = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", url)
    return m.group(1) if m else None

def _get_gspread_client():
    creds_path = os.getenv("GOOGLE_SERVICE_ACCOUNT_FILE", "service_account.json")
    # Also check next to the executable
    if not os.path.isfile(creds_path):
        alt = os.path.join(_base_dir, "service_account.json")
        if os.path.isfile(alt):
            creds_path = alt
        else:
            return None
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_file(creds_path, scopes=scopes)
    return gspread.authorize(creds)

def _detect_headers(ws) -> list[str]:
    """Read the first row to detect existing column headers."""
    row1 = ws.row_values(1)
    return [h.strip() for h in row1 if h.strip()]

def _ensure_headers(ws):
    """Write default headers if row 1 is empty."""
    existing = ws.row_values(1)
    if not any(cell.strip() for cell in existing):
        ws.update([SHEET_HEADERS], "A1")

def _map_record_to_row(record: dict, headers: list[str]) -> list:
    """Map a record dict to a row list matching the sheet's header order."""
    # Normalize header names for matching
    field_map = {
        "timestamp": record["timestamp"],
        "expense/income": record["expense_income"],
        "date": record["date"],
        "month": record["month"],
        "amount": record["amount"],
        "type of expense": record["type_of_expense"],
        "typeofexpense": record["type_of_expense"],
        "expense type": record["type_of_expense"],
        "category": record["type_of_expense"],
        "description": record["description"],
        "desc": record["description"],
        "notes": record["description"],
        "store": record["description"].split(" — ")[0] if " — " in record["description"] else record["description"],
    }

    row = []
    for h in headers:
        key = h.lower().strip()
        row.append(field_map.get(key, ""))
    return row

def append_to_sheets(spreadsheet_id: str, record: dict) -> list[str]:
    """Append one record to the Google Sheet. Returns the headers used."""
    gc = _get_gspread_client()
    if gc is None:
        raise RuntimeError("Service account file not found")

    sh = gc.open_by_key(spreadsheet_id)
    ws = sh.sheet1

    headers = _detect_headers(ws)
    if not headers:
        _ensure_headers(ws)
        headers = SHEET_HEADERS

    row = _map_record_to_row(record, headers)
    ws.append_row(row, value_input_option="USER_ENTERED")
    return headers

# Excel helpers
def build_single_receipt_excel(record: dict, filepath: str):
    """Create a new Excel file with one receipt record."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Receipt"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="7C52B8", end_color="7C52B8", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, header in enumerate(SHEET_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    values = [
        record["timestamp"], record["expense_income"], record["date"],
        record["month"], record["amount"], record["type_of_expense"],
        record["description"],
    ]
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=2, column=col, value=val)
        cell.border = thin_border
        if col == 5:
            cell.number_format = "$#,##0.00"

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = 20

    wb.save(filepath)


# Routes
@app.route("/")
def index():
    return render_template("index.html")

@app.route("/parse", methods=["POST"])
def parse():
    """Upload one receipt image → parse it → return the record."""
    if "file" not in request.files:
        return jsonify({"error": "No file provided"}), 400

    file = request.files["file"]
    if not file or not file.filename or not allowed_file(file.filename):
        return jsonify({"error": "Invalid file type"}), 400

    filename = secure_filename(file.filename)
    unique = f"{uuid.uuid4().hex}_{filename}"
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], unique)
    file.save(filepath)

    try:
        parsed = parse_receipt(filepath)
        record = receipt_to_record(parsed)
        return jsonify({"record": record, "raw": parsed})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

@app.route("/submit", methods=["POST"])
def submit():
    """Take a parsed record and send it to the chosen destination."""
    data = request.get_json()
    if not data or "record" not in data:
        return jsonify({"error": "No record provided"}), 400

    record = data["record"]
    dest = data.get("destination", "sheets")  # "sheets" or "excel"
    action = data.get("action", "append")     # "append" or "new"

    if dest == "sheets" and action == "append":
        if not GSHEETS_AVAILABLE:
            return jsonify({"error": "Google Sheets libraries not installed"}), 400
        if not sheets_state["connected"]:
            return jsonify({"error": "No Google Sheet connected"}), 400
        try:
            headers = append_to_sheets(sheets_state["spreadsheet_id"], record)
            return jsonify({"message": "Row added to Google Sheets", "headers": headers})
        except Exception as e:
            return jsonify({"error": str(e)}), 500

    elif dest == "excel" or action == "new":
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"receipt_{ts}.xlsx"
        filepath = os.path.join(app.config["EXPORT_FOLDER"], filename)
        build_single_receipt_excel(record, filepath)
        return jsonify({"message": "Excel file created", "download": f"/download/{filename}"})

    return jsonify({"error": "Invalid destination/action"}), 400

@app.route("/download/<filename>")
def download(filename):
    safe = secure_filename(filename)
    filepath = os.path.join(app.config["EXPORT_FOLDER"], safe)
    if not os.path.isfile(filepath):
        return jsonify({"error": "File not found"}), 404
    return send_file(filepath, as_attachment=True, download_name=safe)

# Google Sheets connection
@app.route("/sheets/connect", methods=["POST"])
def sheets_connect():
    if not GSHEETS_AVAILABLE:
        return jsonify({"error": "Google Sheets libraries not installed. "
                        "Run: pip install gspread google-auth"}), 400

    data = request.get_json()
    url = (data or {}).get("url", "").strip()
    if not url:
        return jsonify({"error": "No URL provided"}), 400

    sid = _extract_spreadsheet_id(url)
    if not sid:
        return jsonify({"error": "Invalid Google Sheets URL"}), 400

    gc = _get_gspread_client()
    if gc is None:
        return jsonify({"error": "service_account.json not found"}), 400

    try:
        sh = gc.open_by_key(sid)
        title = sh.title
    except Exception as e:
        return jsonify({"error": f"Cannot access sheet: {e}"}), 400

    sheets_state["connected"] = True
    sheets_state["spreadsheet_id"] = sid
    sheets_state["sheet_url"] = url

    return jsonify({"title": title})

@app.route("/sheets/status")
def sheets_status():
    return jsonify(sheets_state)

@app.route("/sheets/disconnect", methods=["POST"])
def sheets_disconnect():
    sheets_state.update({"connected": False, "spreadsheet_id": None, "sheet_url": None})
    return jsonify({"message": "Disconnected"})

# PWA manifest
@app.route("/manifest.json")
def manifest():
    m = {
        "name": "Receipt Reader",
        "short_name": "Receipts",
        "start_url": "/",
        "display": "standalone",
        "background_color": "#f8f0ff",
        "theme_color": "#9b72cf",
        "icons": [
            {"src": "/static/icon-192.png", "sizes": "192x192", "type": "image/png"},
            {"src": "/static/icon-512.png", "sizes": "512x512", "type": "image/png"},
        ],
    }
    return jsonify(m)

# Entry point
if __name__ == "__main__":
    port = int(os.getenv("PORT", 5000))
    is_exe = getattr(sys, "frozen", False)

    if is_exe:
        try:
            import webview
            webview.create_window("Receipt Reader", app, width=420, height=720)
            webview.start()
        except ImportError:
            import webbrowser, threading
            threading.Timer(1.0, lambda: webbrowser.open(f"http://localhost:{port}")).start()
            app.run(port=port)
    else:
        app.run(debug=True, port=port)
